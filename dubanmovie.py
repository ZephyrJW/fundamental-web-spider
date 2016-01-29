#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')



#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def movie_spider(movie_tag):
    page_num=0;
    movie_list=[]
    try_times=0
    
    while(1):
        #url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        url='http://www.douban.com/tag/'+urllib.quote(movie_tag)+'/movie?start='+str(page_num*15)
        time.sleep(np.random.rand()*5)
        
        #Last Version
        try:
            req = urllib2.Request(url, headers=hds[page_num%len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text=str(source_code)   
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue
  
        ##Previous Version, IP is easy to be Forbidden
        #source_code = requests.get(url) 
        #plain_text = source_code.text  
        
        soup = BeautifulSoup(plain_text)
        list_soup = soup.find('div', {'class': 'mod movie-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for movie_info in list_soup.findAll('dd'):
            title = movie_info.find('a', {'class':'title'}).string.strip()
            desc = movie_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            movie_url = movie_info.find('a', {'class':'title'}).get('href')
            
            try:
                tag_info = '/'.join(desc_list[0:-5]) #str.join("a","b","c") -> a+str+b+str+c eg: a/b/c
            except:
                tag_info ='标签： 暂无'
            try:
                pub_info = '上映时间： ' + desc_list[-5]
            except:
                pub_info = '上映时间： 暂无'
            try:
                director_info = '导演: ' +  desc_list[-4]
            except:
                director_info = '导演: 暂无' 
            try:
                actor_info = '主演: ' + '/'.join(desc_list[-3:-1])
            except:
                actor_info = '主演: 暂无'
            try:
                rating = movie_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                #people_num = book_info.findAll('span')[2].string.strip()
                people_num = get_people_num(movie_url)
                people_num = people_num.strip('人评价')
            except:
                people_num ='0'
            
            movie_list.append([title,tag_info,rating,people_num,pub_info,director_info,actor_info])
            try_times=0 #set 0 when got valid information
        page_num+=1
        print 'Downloading Information From Page %d' % page_num
    return movie_list


def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib2.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib2.urlopen(req).read()
        plain_text=str(source_code)   
    except (urllib2.HTTPError, urllib2.URLError), e:
        print e
    soup = BeautifulSoup(plain_text)
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[0].string.strip()
    return people_num


def do_spider(movie_tag_lists):
    movie_lists=[]
    for movie_tag in movie_tag_lists:
        movie_list=movie_spider(movie_tag)
        movie_list=sorted(movie_list,key=lambda x:x[1],reverse=True) # what's x[1]? look line 88, it's rating!
        movie_lists.append(movie_list)
    return movie_lists


def print_movie_lists_excel(movie_lists,movie_tag_lists):
    wb=Workbook(optimized_write=True)
    ws=[]
    for i in range(len(movie_tag_lists)):
        ws.append(wb.create_sheet(title=movie_tag_lists[i].decode())) #utf8->unicode
    for i in range(len(movie_tag_lists)): 
        ws[i].append(['序号','影名','标签','评分','评价人数','上映年份','导演','主演'])
        count=1
        for ml in movie_lists[i]:
            ws[i].append([count,ml[0],ml[1],float(ml[2]),ml[3],ml[4],ml[5],ml[6]])
            count+=1
    save_path='movie_list'
    for i in range(len(movie_tag_lists)):
        save_path+=('-'+movie_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path)




if __name__=='__main__':
    movie_tag_lists = ['杨德昌','李安'] #test

    movie_lists=do_spider(movie_tag_lists)
    print_movie_lists_excel(movie_lists,movie_tag_lists)
